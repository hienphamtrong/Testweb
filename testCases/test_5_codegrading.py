from testCases import wait_click
from utilities.customLogger import LogGen
import testCases.test_1_home as br_st
import testCases.variables as vr
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import random
import openpyxl

values_lable = "//*[contains(text(),'values')]"
grades_lable = "//*[contains(text(),'grades')]"
check_box_xpath = "(//*[@class='mat-checkbox-inner-container'])"
graphs_tagname = "app-dl-grading-graph"
setting_button = "//*[contains(text(),'settings')]"
pause_button = "//*[contains(text(),'pause')]"
play_button = "//*[contains(text(),'play_arrow')]"
ref_box = "//*[@placeholder='Refresh Period (ms)']"
sample_box = "//*[@placeholder='Number of Samples']"
menu_button_xpath = "//mat-icon[contains(text(),'menu')]"
codegrading_item_xpath = "//*[normalize-space()='Code Grading Monitoring']"
values_lable_xpath_hide = "//*[*[contains(.,'Values')] and @aria-expanded='false']"
grades_lable_xpath_hide = "//*[*[contains(.,'Grades')] and @aria-expanded='false']"
data_wb = openpyxl.load_workbook('./TestData/data_code_grading.xlsx')
data_xpath = "//*[contains(text(),'"
logger = LogGen.loggen()
flag = True
time_out = 5
no_graphs_no_checkbox = False


def data_content_verify(driver):
    global flag
    logger.info("**********Verify label information!*********\n")
    try:
        max_row = data_wb['content'].max_row
        for row in range(1, max_row + 1):
            read_data = data_wb['content'].cell(row, 1).value
            full_data_xpath = data_xpath + str(read_data) + "')]"
            try:
                driver.find_element_by_xpath(full_data_xpath)
                logger.info("Label- Expected: " + str(read_data) + "  - Actual: " + str(driver.find_element_by_xpath(full_data_xpath).text))
            except Exception:
                flag = False
                logger.error(str(read_data) + " not found")
    except Exception:
        logger.error("**********Label test is failed!*********\n")
    logger.info("**********Label test is completed!*********\n")


def data_checkbox_verify(driver):
    global flag
    logger.info("**********Verify Checkbox Name!*********\n")
    try:
        max_row = data_wb['checkbox_name'].max_row
        for row in range(1, max_row + 1):
            read_data = data_wb['checkbox_name'].cell(row, 1).value
            full_data_xpath = data_xpath + str(read_data) + "')]"
            check_box_name = driver.find_element_by_xpath(
                "(//*[@class='mat-checkbox-label'])" + '[' + str(row) + ']').text
            try:
                driver.find_element_by_xpath(full_data_xpath)
                logger.info("Expected: " + str(read_data) + " - Actual: " + str(check_box_name))
            except Exception:
                flag = False
                logger.error("Expected: " + str(read_data) + " - Actual: " + str(check_box_name))

    except Exception:
        logger.error("**********Checkbox Name test is failed!*********\n")
    logger.info("**********Checkbox Name test is completed!*********\n")


def code_grading_verify(driver):
    global flag, num_graphs_show_in_grades, num_graphs_show_in_values, grading
    try:
        driver.find_element_by_xpath(menu_button_xpath).click()
        wait_click.click_until_interactable(driver.find_element_by_xpath(codegrading_item_xpath))
    except Exception:
        flag = False
        logger.error("Error Click - Codegrading")

    try:
        logger.info("**********Verify Grades/Values!***************\n")
        try:
            if WebDriverWait(driver, time_out).until(EC.presence_of_element_located((By.TAG_NAME, graphs_tagname))):
                grading = True
                logger.info("ISO/IEC 15416 Grading is enabled")
        except Exception:
            grading = False
            logger.info("No Grading is enabled")
            pass

        WebDriverWait(driver, time_out).until(EC.presence_of_element_located((By.XPATH, values_lable_xpath_hide)))
        num_graphs_show_in_grades = len(driver.find_elements_by_tag_name(graphs_tagname))
        driver.find_element_by_xpath(values_lable).click()

        WebDriverWait(driver, time_out).until(EC.presence_of_element_located((By.XPATH, grades_lable_xpath_hide)))
        num_graphs_show_in_values = len(driver.find_elements_by_tag_name(graphs_tagname))
        if num_graphs_show_in_values != num_graphs_show_in_grades:
            logger.error("Graphs Show in Values: " + str(num_graphs_show_in_values))
            logger.error("Graphs Show in Grades: " + str(num_graphs_show_in_grades))
            flag = False
        driver.find_element_by_xpath(grades_lable).click()
    except Exception:
        flag = False
        logger.error("Graph plotting Values/Grades is not worked!")
    logger.info("**********Grades/Values test is completed!***************\n")

    driver.find_element_by_xpath(setting_button).click()
    if grading != True:
        data_content_verify(vr.driver)
    else:
        data_content_verify(vr.driver)
        data_checkbox_verify(vr.driver)

    try:
        logger.info("**********Verify No. of graphs/checkbox!***************\n")
        num_check_box = len(driver.find_elements_by_xpath(check_box_xpath))
        if grading == True:
            logger.info("No. of graphs Expected: 9 " + " - Actual: " + str(num_graphs_show_in_grades))
            logger.info("No. of checkbox Expected: 9 " + " - Actual: " + str(num_check_box))
        else:
            logger.info("No. of graphs Expected: 0 " + " - Actual: " + str(num_graphs_show_in_grades))
            logger.info("No. of checkbox Expected: 0 " + " - Actual: " + str(num_check_box))

        if num_check_box != num_graphs_show_in_grades:
            flag = False
            logger.error("No. of checkbox are different No. of graphs")
    except Exception:
        flag = False
    logger.info("**********No. of graphs/checkbox test is completed!***************\n")

    logger.info("**********Verify Graph Options!***************\n")
    try:
        driver.find_element_by_xpath(sample_box).clear()
        random_input = random.randint(1, 500)
        driver.find_element_by_xpath(sample_box).send_keys(random_input)
        driver.find_element_by_xpath(pause_button).click()
        driver.find_element_by_xpath(play_button).click()
        # if random_input != driver.find_element_by_xpath(sample_box).text:
        #     logger.error("No. of Samples Expected: " + str(random_input) + " - Actual: " + driver.find_element_by_xpath(sample_box).text)
        # else:
        #     logger.info("No. of Samples Expected: " + str(random_input) + " - Actual: " + driver.find_element_by_xpath(sample_box).text)

        driver.find_element_by_xpath(ref_box).clear()
        random_input = random.randint(1, 500)
        driver.find_element_by_xpath(ref_box).send_keys(random_input)
        driver.find_element_by_xpath(pause_button).click()
        driver.find_element_by_xpath(play_button).click()

        # if random_input != driver.find_element_by_xpath(ref_box).text:
        #     logger.error("No. of Samples Expected: " + str(random_input) + " - Actual: " + str(driver.find_element_by_xpath(ref_box).text))
        # else:
        #     logger.info("No. of Samples Expected: " + str(random_input) + " - Actual: " + str(driver.find_element_by_xpath(ref_box).text))
    except Exception:
        flag = False
        logger.error("**********Verify Graph Options is failed!***************\n")
    logger.info("**********Verify Graph Options test is completed!***************\n")
    if num_check_box != 0:
        try:
            logger.info("**********Verify Checkbox function!***************\n")
            for ttt in range(1, num_check_box + 1):
                driver.find_element_by_xpath(check_box_xpath + "[" + str(ttt) + "]").click()

            for ttt in range(1, num_check_box + 1):
                driver.find_element_by_xpath(check_box_xpath + "[" + str(ttt) + "]").click()

            if grading == True:
                for t in range(1, random.randint(2, 6)):  # random number of turns
                    for tt in range(1, random.randint(2, num_check_box) + 1):  # random number of boxes to click
                        driver.find_element_by_xpath(
                            check_box_xpath + "[" + str(random.randint(1, num_check_box)) + "]").click()  # random index box
                    num_checked_boxes = len(driver.find_elements_by_xpath("//*[@aria-checked='true']"))
                    graphs_util_click = len(driver.find_elements_by_tag_name(graphs_tagname))
                    if num_checked_boxes != graphs_util_click:
                        flag = False
                        logger.error("No. of checked box : " + str(num_checked_boxes))
                        logger.error("No. of graphs Expected: " + str(num_checked_boxes) + " - Actual: " + str(graphs_util_click))
                    else:
                        logger.info("No. of checked box : " + str(num_checked_boxes))
                        logger.info("No. of graphs Expected: " + str(num_checked_boxes) + " - Actual: " + str(graphs_util_click))
            driver.find_element_by_xpath(setting_button).click()

        except Exception:
            flag = False
            logger.error("**********Checkbox function test is failed!***************\n")
        logger.info("**********Checkbox function test is completed!***************\n")


def test_code_grading(browser):
    global flag
    if vr.driver is None:
        br_st.browser_setup(browser)
    logger.info("**********Test_005_Code_Grading_Monitoring***************\n")
    code_grading_verify(vr.driver)
    if flag == False:
        assert False
