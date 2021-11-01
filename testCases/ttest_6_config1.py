import openpyxl
from testCases import wait_click
from utilities.customLogger import LogGen
import testCases.test_1_home as br_st
import testCases.variables as vr
import random
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pytest
import time

logger = LogGen.loggen()
alarm_xpath = "//*[normalize-space()='"
job_wb = openpyxl.load_workbook('./TestData/data_config.xlsx')
menu_button_xpath = "//mat-icon[contains(text(),'menu')]"
config_item_xpath = "//*[normalize-space()='Configurator']"
default_config_xpath = "//*[normalize-space()='Default']"
present_default_config_xpath = "//*[normalize-space()='PresentationDefault']"
job_xpath = "//*[contains(text(),'"
xpath_1 = "//*[*[normalize-space()='"
xpath_2 = "'] and *//*[normalize-space()='done' and @class='material-icons selected-icon']]"
xpath_3 = "'] and *//*[normalize-space()='done_all' and @class='material-icons selected-icon']]"
running_row = "//*[*[text()=' Running ']]"
startup_row = "//*[*[text()=' Startup ']]"

load_xpath = "(//*[text()='done'])"
load_all_xpath = "(//*[text()='done_all'])"
save_config_check_xpath = "(//*[contains(@class,'column')][1])"

flag = True


def replace_data_config(data):
    data_1 = data.replace("\n", "")
    data = data_1.replace("donedone_all", "")
    return data


def data_config_verify(driver):
    global flag, data_11, data_22, max_row, check_load, num_save_config
    try:
        driver.find_element_by_xpath(menu_button_xpath).click()
        wait_click.click_until_interactable(driver.find_element_by_xpath(config_item_xpath))
    except Exception:
        flag = False
        logger.error("Error Click - Configurator")

    try:
        logger.info("Verify data, current Running/Startup config")
        driver.find_element_by_xpath(default_config_xpath)
        driver.find_element_by_xpath(present_default_config_xpath)
        max_row = job_wb['Sheet1'].max_row
        for row in range(1, max_row + 1):
            read_data = job_wb['Sheet1'].cell(row, 1).value
            full_data_xpath = job_xpath + str(read_data) + "')]"
            driver.find_element_by_xpath(full_data_xpath)
            try:
                if driver.find_element_by_xpath(xpath_1 + str(read_data) + xpath_2) != None:
                    data_1 = driver.find_element_by_xpath(xpath_1 + str(read_data) + xpath_2).text
                    data_11 = replace_data_config(data_1)
                check_load = driver.find_element_by_xpath(running_row).text
                if not (data_11 in check_load):
                    flag = False
                    logger.error("Expected: Running " + data_11 + " Actual: Running " + str(check_load))

            except Exception:
                pass

            try:
                if driver.find_element_by_xpath(xpath_1 + str(read_data) + xpath_3) != None:
                    data_2 = driver.find_element_by_xpath(xpath_1 + str(read_data) + xpath_3).text
                    data_22 = replace_data_config(data_2)
                check_load = driver.find_element_by_xpath(startup_row).text
                if not (data_22 in check_load):
                    flag = False
                    logger.error("Expected: Startup " + data_22 + " Actual: Startup " + str(check_load))
            except Exception:
                pass
    except Exception:
        flag = False
    logger.info("Verify data, current Running/Startup config is completed!")

    num_save_config = len(driver.find_elements_by_xpath(save_config_check_xpath))
    logger.info("No. of saved Configuration Expected: " + str(max_row) + " - Actual: " + str(num_save_config - 4))
    # sequence_check
    logger.info("Verify Load button in Saved Configurations!")
    for run in range(1, max_row + 1):
        try:
            current_data_running = driver.find_element_by_xpath(running_row).text
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, load_xpath + "[" + str(run + 2) + "]")))
            try:
                driver.find_element_by_xpath(load_xpath + "[" + str(run + 2) + "]").click()
            except Exception:
                driver.find_element_by_xpath(load_xpath + "[" + str(run + 2) + "]").click()
                pass
            wait_click.is_green(driver.find_element_by_xpath(load_xpath + "[" + str(run + 2) + "]"))
            wait_click.wait_data_change(driver.find_element_by_xpath(running_row), current_data_running)
        except Exception:
            flag = False
            logger.error("Failed in Click")

        read_config = driver.find_element_by_xpath(save_config_check_xpath + "[" + str(run + 4) + "]").text
        check_load = driver.find_element_by_xpath(running_row).text
        if not (read_config in check_load):
            flag = False
            logger.error(
                "Expected: Running " + str(read_config) + " - Actual: " + str(check_load.replace("\n", " ")))


def test_configurator(browser):
    global flag
    if vr.driver is None:
        br_st.browser_setup(browser)
    logger.info("**********Test_006_ConfiguratorPage***************\n")
    data_config_verify(vr.driver)
    # vr.driver.quit()
    if flag == False:
        assert False


