import openpyxl
from testCases import wait_click
from utilities.customLogger import LogGen
import testCases.test_1_home as br_st
import testCases.variables as vr
import time

menu_button_xpath = "//mat-icon[contains(text(),'menu')]"
monitor_item_xpath = "//*[normalize-space()='Statistics']"
logger = LogGen.loggen()
flag = True
data_wb = openpyxl.load_workbook('./TestData/data_statistics.xlsx')
info_xpath = "*[normalize-space()='"


def data_statistics_verify(driver):
    global flag, read_again
    try:
        driver.find_element_by_xpath(menu_button_xpath).click()
        wait_click.click_until_interactable(driver.find_element_by_xpath(monitor_item_xpath))
    except Exception:
        flag = False
        logger.error("Error Click - Statistics")
        # check data
    try:
        max_row = data_wb['Sheet1'].max_row
        for row in range(1, max_row + 1):
            read_data_col_1 = data_wb['Sheet1'].cell(row, 1).value
            try:
                full_xpath = "//" + info_xpath + str(read_data_col_1) + "']"
                driver.find_element_by_xpath(full_xpath)
                # data_display = driver.find_element_by_xpath("//*[" + info_xpath + str(read_data_col_1) + "']]").text
            except Exception:
                logger.error("Expected: " + str(read_data_col_1) + "  -Actual: not found")
        # check value before refresh
        good_read_val_bf = driver.find_element_by_xpath("//*[" + info_xpath + "Good Reads']]").text
        no_read_val_bf = driver.find_element_by_xpath("//*[" + info_xpath + "No Reads']]").text
        good_read_val_bf = good_read_val_bf.splitlines()[1]
        no_read_val_bf = no_read_val_bf.splitlines()[1]
        logger.info("Good Reads value before reset: " + good_read_val_bf)
        logger.info("No Reads value before reset: " + no_read_val_bf)
        # click refresh button
        refresh_1_xpath = "(//*[contains(text(),'refresh')])[1]"
        refresh_2_xpath = "(//*[contains(text(),'refresh')])[2]"
        wait_click.click_until_interactable(driver.find_element_by_xpath(refresh_1_xpath))
        wait_click.click_until_interactable(driver.find_element_by_xpath(refresh_2_xpath))
        # check value before refresh 1s
        time.sleep(1)
        good_read_val_af = driver.find_element_by_xpath("//*[" + info_xpath + "Good Reads']]").text
        no_read_val_af = driver.find_element_by_xpath("//*[" + info_xpath + "No Reads']]").text
        good_read_val_af = good_read_val_af.splitlines()[1]
        no_read_val_af = no_read_val_af.splitlines()[1]
        # read_again = 0
        # while (int(good_read_val_af) >= int(good_read_val_bf) or int(no_read_val_af) >= int(no_read_val_bf)) \
        #         and read_again < 3:
        #
        #     good_read_val_af = driver.find_element_by_xpath("//*[" + info_xpath + "Good Reads']]").text
        #     no_read_val_af = driver.find_element_by_xpath("//*[" + info_xpath + "No Reads']]").text
        #     good_read_val_af = good_read_val_af.splitlines()[1]
        #     no_read_val_af = no_read_val_af.splitlines()[1]
        #     time.sleep(0.5)

        logger.info("Good Reads value after reset: " + good_read_val_af)
        logger.info("No Reads value after reset: " + no_read_val_af)

        if int(good_read_val_af) <= int(good_read_val_bf):
            logger.info("Refresh button is worked!\n")
        else:
            flag = False
            logger.error("Refresh button is not worked!\n")
    except Exception:
        flag = False
        logger.error("Error in verify data statistics")


def test_statistics(browser):
    global flag
    if vr.driver is None:
        br_st.browser_setup(browser)
    logger.info("**********Test_003_StatisticsPage***************\n")
    data_statistics_verify(vr.driver)
    if flag == False:
        assert False
