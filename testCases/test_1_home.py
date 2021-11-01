import openpyxl

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from utilities.customLogger import LogGen
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from testCases import wait_click
import testCases.variables as vr

baseURL = "http://matrix-user:D@talogic1972@127.0.0.1"
# baseURL = "http://127.0.0.1"
logger = LogGen.loggen()
menuitem_xpath = "(//*[@role='menuitem'])"
menubutton_xpath = "//mat-icon[contains(text(),'menu')]"
flag = True

wb = openpyxl.load_workbook("./TestData/data_device.xlsx")
chromeOptions = Options()
chromeOptions.add_experimental_option("prefs", {"download.default_directory": "C:\\Users\\24H\\PycharmProjects\\pytest_web\\Download",
                                                "safebrowsing.enabled": False})


def browser_setup(browser):
    if browser == 'chrome':
        vr.driver = webdriver.Chrome(
            executable_path="C:\\Users\\24H\\PycharmProjects\\pytest_web\\drivers\\chromedriver.exe",
            chrome_options=chromeOptions)
    elif browser == 'firefox':
        vr.driver = webdriver.Firefox(executable_path="./drivers/geckodriver.exe")
    logger.info("Browser Name: " + vr.driver.capabilities['browserName'])
    logger.info("Version: " + vr.driver.capabilities['browserVersion'])
    vr.driver.maximize_window()
    try:
        vr.driver.get(baseURL)
    except Exception:
        logger.error("Error URL")


def menu_button_verify(driver):
    global flag
    logger.info("**********Verify Menu button!**************")
    # driver.get("http://matrix-user:D@talogic1972@127.0.0.1")
    try:
        driver.find_element_by_xpath(menubutton_xpath).click()
        num_menuitem = len(driver.find_elements_by_xpath(menuitem_xpath))
        for item in range(1, num_menuitem + 1):
            menu_item = driver.find_element_by_xpath(menuitem_xpath + "[" + str(item) + "]")
            wait_click.click_until_interactable(menu_item)

            WebDriverWait(driver, 10).until(EC.url_to_be(driver.current_url))
            driver.find_element_by_xpath(menubutton_xpath).click()
        wait_click.click_any_where(driver, 1)
    except Exception:
        flag = False
        logger.error("Clicking the menuitem failed")
    logger.info("**********Menu button test is completed!**************")


def data_homepage_verify(driver):
    global flag
    logger.info("**********Verify Information!**************")
    try:
        max_row = wb['Sheet2'].max_row
        info_xpath = "*[normalize-space()='"
        for row in range(1, max_row + 1):
            read_data_col_1 = wb['Sheet1'].cell(row, 1).value
            read_data_col_2 = wb['Sheet1'].cell(row, 2).value
            try:
                full_info_xpath = "//*[" + info_xpath + str(read_data_col_1) + "'] and " + info_xpath + str(
                    read_data_col_2) + "']]"
                driver.find_element_by_xpath(full_info_xpath)
            except Exception:
                flag = False
                try:
                    driver.find_element_by_xpath("//*[" + info_xpath + str(read_data_col_1) + "']]")
                    data_display = driver.find_element_by_xpath("//*[" + info_xpath + str(read_data_col_1) + "']]").text
                    data_replace = data_display.replace("\n", " ")
                    logger.error("Expected: " + str(read_data_col_1) + ' ' + str(read_data_col_2) + "  -Actual: " + str(
                        data_replace))
                except Exception:
                    logger.error("Expected: " + str(read_data_col_1) + "  -Actual: not found")
        if flag == True:
            logger.info("**********Information test is passed!*********\n")
    except Exception:
        logger.error("**********Information test is failed!*********\n")
    logger.info("**********Information test is completed!*********\n")


def test_homepage(browser):
    global flag
    browser_setup(browser)
    logger.info("**********Test_001_HomePage***************\n")
    data_homepage_verify(vr.driver)
    menu_button_verify(vr.driver)
    if flag == False:
        assert False

# test_homepage('chrome')
