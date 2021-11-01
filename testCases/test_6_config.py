import openpyxl
from testCases import wait_click
from utilities.customLogger import LogGen
import testCases.test_1_home as br_st
import testCases.variables as vr
import random
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time

logger = LogGen.loggen()

job_wb = openpyxl.load_workbook('./TestData/data_config.xlsx')
menu_button_xpath = "//mat-icon[contains(text(),'menu')]"
config_item_xpath = "//*[normalize-space()='Configurator']"
default_config_xpath = "//*[normalize-space()='Default']"
present_default_config_xpath = "//*[normalize-space()='PresentationDefault']"
job_xpath = "//*[contains(text(),'"

running_row = "//*[*[text()=' Running ']]"
startup_row = "//*[*[text()=' Startup ']]"
load_ena_xpath = "//*[normalize-space()='done' and @class='material-icons selected-icon']"
load_xpath = "(//*[text()='done'])"
load_all_xpath = "(//*[text()='done_all'])"
save_config_check_xpath = "(//*[contains(@class,'column')][1])"

selected_load = "//*[*/*/*/*[normalize-space()='done' and @class='material-icons selected-icon']]"
selected_all_load = "//*[*/*/*/*[normalize-space()='done_all' and @class='material-icons selected-icon']]"

flag = True


def replace_data_config(data):
    data_1 = data.replace("\n", "")
    data = data_1.replace("donedone_all", "")
    return data


def data_config_verify(driver):
    global flag, max_row
    try:
        driver.find_element_by_xpath(menu_button_xpath).click()
        wait_click.click_until_interactable(driver.find_element_by_xpath(config_item_xpath))
        WebDriverWait(driver, 10).until(EC.url_to_be(driver.current_url))

    except Exception:
        flag = False
        logger.error("Error Click - Configurator\n")

    try:
        logger.info("Verify current Configuration\n")
        running_current = replace_data_config(driver.find_element_by_xpath(selected_load).text)
        startup_current = replace_data_config(driver.find_element_by_xpath(selected_all_load).text)
        running_config = driver.find_element_by_xpath(running_row).text
        startup_config = driver.find_element_by_xpath(startup_row).text
        if not (running_current in running_config):
            flag = False
            logger.error("Expected: Running " + running_current + " Actual:  " + running_config.replace("\n", " "))
        else:
            logger.info("Expected: Running " + running_current + " Actual:  " + running_config.replace("\n", " "))

        if not (startup_current in startup_config):
            flag = False
            logger.error("Expected: Running " + startup_current + " Actual:  " + startup_config.replace("\n", " "))
        else:
            logger.info("Expected: Running " + startup_current + " Actual:  " + startup_config.replace("\n", " "))

    except Exception:
        flag = False
        logger.error("Current Configuration is Failed!\n")
    logger.info("Current Configuration test is completed!\n")

    logger.info("Verify Default Configurations\n")
    try:
        driver.find_element_by_xpath(default_config_xpath)
        driver.find_element_by_xpath(present_default_config_xpath)
    except Exception:
        flag = False
        logger.error("Default Configurations test is Failed!\n")
    logger.info("Default Configurations test is completed!\n")

    logger.info("Verify Saved Configurations\n")
    try:
        num_save_config = len(driver.find_elements_by_xpath(save_config_check_xpath))
        max_row = job_wb['Sheet1'].max_row
        if max_row != (num_save_config-4):
            flag = False
            logger.error("No. of saved Configuration Expected: " + str(max_row) + " - Actual: " + str(num_save_config - 4))
        else:
            logger.info("No. of saved Configuration Expected: " + str(max_row) + " - Actual: " + str(num_save_config - 4))

        for row in range(1, max_row + 1):
            read_data = job_wb['Sheet1'].cell(row, 1).value
            full_data_xpath = job_xpath + str(read_data) + "')]"
            try:
                driver.find_element_by_xpath(full_data_xpath)
            except Exception:
                flag = False
                logger.error("Expected: " + str(read_data) + "  -Actual: not found")
    except Exception:
        flag = False
    logger.info("Saved Configurations test is completed!\n")

    logger.info("Verify Load button!\n")
    try:
        for run in range(1, max_row + 3):
            try:
                current_data_running = driver.find_element_by_xpath(running_row).text
                startup_data_bf = driver.find_element_by_xpath(startup_row).text
                driver.find_element_by_xpath(load_xpath + "[" + str(run) + "]").click()
            except Exception:
                current_data_running = driver.find_element_by_xpath(running_row).text
                startup_data_bf = driver.find_element_by_xpath(startup_row).text
                driver.find_element_by_xpath(load_xpath + "[" + str(run) + "]").click()
                pass
            if run == 2:
                read_config = "[Temp]"
                wait_click.wait_data_change(driver.find_element_by_xpath(running_row), current_data_running)
                WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.XPATH, load_ena_xpath)))
            else:
                wait_click.is_green(driver.find_element_by_xpath(load_xpath + "[" + str(run) + "]"))
                wait_click.wait_data_change(driver.find_element_by_xpath(running_row), current_data_running)
                try:
                    read_config = driver.find_element_by_xpath(selected_load).text
                except Exception:
                    read_config = driver.find_element_by_xpath(selected_load).text
                    pass
            try:
                check_load = driver.find_element_by_xpath(running_row).text
                startup_data_af = driver.find_element_by_xpath(startup_row).text
            except Exception:
                check_load = driver.find_element_by_xpath(running_row).text
                pass
            if not (replace_data_config(read_config) in check_load):
                flag = False
                logger.error(
                    "Expected: Running " + str(replace_data_config(read_config)) + " - Actual: " + str(check_load.replace("\n", " ")) +
                    "\n                                   Expected: " + str(startup_data_bf.replace("\n", " ")) + " - Actual: " + str(
                        startup_data_af.replace("\n", " ")))

            else:
                logger.info(
                    "Expected: Running " + str(replace_data_config(read_config)) + " - Actual: " + str(check_load.replace("\n", " ")) +
                    "\n                                   Expected: " + str(startup_data_bf.replace("\n", " ")) + " - Actual: " + str(
                        startup_data_af.replace("\n", " ")))
    except Exception:
        flag = False
        logger.error("Load button test is failed!\n")
    logger.info("Load button test is completed!\n")

    logger.info("Verify Load and Set Startup button!\n")
    try:
        for run in range(1, max_row + 3):
            if run != 2:
                try:
                    current_data_running = driver.find_element_by_xpath(running_row).text
                    current_data_startup = driver.find_element_by_xpath(startup_row).text
                    driver.find_element_by_xpath(load_all_xpath + "[" + str(run) + "]").click()
                except Exception:
                    driver.find_element_by_xpath(load_all_xpath + "[" + str(run) + "]").click()
                    current_data_running = driver.find_element_by_xpath(running_row).text
                    current_data_startup = driver.find_element_by_xpath(startup_row).text
                    pass

                wait_click.is_green(driver.find_element_by_xpath(load_xpath + "[" + str(run) + "]"))
                wait_click.is_green(driver.find_element_by_xpath(load_all_xpath + "[" + str(run) + "]"))

                wait_click.wait_data_change(driver.find_element_by_xpath(running_row), current_data_running)
                wait_click.wait_data_change(driver.find_element_by_xpath(startup_row), current_data_startup)
                try:
                    read_config = driver.find_element_by_xpath(selected_all_load).text
                    check_load = driver.find_element_by_xpath(running_row).text
                    check_startup = driver.find_element_by_xpath(startup_row).text
                except Exception:
                    read_config = driver.find_element_by_xpath(selected_all_load).text
                    check_load = driver.find_element_by_xpath(running_row).text
                    check_startup = driver.find_element_by_xpath(startup_row).text
                    pass

                if (not (replace_data_config(read_config) in check_load)) and \
                        (not (replace_data_config(read_config)) in check_startup):
                    flag = False
                    logger.error(
                        "Expected: Running " + str(replace_data_config(read_config)) + " Startup  " + str(replace_data_config(read_config)) +
                        "\n                                   Actual:   " + str(check_load.replace("\n", " ")) + " " + str(check_startup.replace("\n", " ")))
                else:
                    logger.info(
                        "Expected: Running " + str(replace_data_config(read_config)) + " Startup  " + str(replace_data_config(read_config)) +
                        "\n                                   Actual:   " + str(check_load.replace("\n", " ")) + " " + str(check_startup.replace("\n", " ")))

    except Exception:
        flag = False
        logger.error("Load and Set Startup button test is failed!\n")
    logger.info("Load and Set Startup button test is completed!\n")


def test_configurator(browser):
    global flag
    if vr.driver is None:
        br_st.browser_setup(browser)
    logger.info("**********Test_006_ConfiguratorPage***************\n")
    data_config_verify(vr.driver)
    vr.driver.quit()
    if flag == False:
        assert False

