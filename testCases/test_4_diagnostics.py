import openpyxl

from testCases import wait_click
from utilities.customLogger import LogGen
import testCases.test_1_home as br_st
import testCases.variables as vr
import time

logger = LogGen.loggen()
alarm_wb = openpyxl.load_workbook('./TestData/Alarm.xlsx')

alarm_xpath = "*[normalize-space()='"
menu_button_xpath = "//mat-icon[contains(text(),'menu')]"
diagnostics_item_xpath = "//*[normalize-space()='Diagnostics']"
row_xpath = "//*[@role ='row']"
red_alarm_xpath = "//*[@role ='row'and */*[@color='red']]"

flag = True


def replace_data_alarm(data):
    data_1 = data.replace("\n", " ")
    data = data_1.replace("●", "")
    return data


def data_diagnostics_verify(driver):
    global flag, alarm_display_data
    try:
        driver.find_element_by_xpath(menu_button_xpath).click()
        wait_click.click_until_interactable(driver.find_element_by_xpath(diagnostics_item_xpath))
    except Exception:
        flag = False
        logger.error("Error Click - Diagnostics")
    try:
        logger.info("**********Verify Alarm information!***************\n")
        max_col = alarm_wb['Sheet1'].max_column
        max_row = alarm_wb['Sheet1'].max_row
        try:
            num_alarm = len(driver.find_elements_by_xpath("//*[@color='green' or @color='red']"))
            if max_row == num_alarm:
                logger.info("No. of Alarm Expected: " + str(max_row) + " - Actual: " + str(num_alarm))
            else:
                flag = False
                logger.error("No. of Alarm Expected: " + str(max_row) + " - Actual: " + str(num_alarm))
        except Exception:
            logger.error("Failed")
            pass

        for row in range(1, max_row + 1):
            read_data_col_1 = alarm_wb['Sheet1'].cell(row, 1).value
            read_data_col_2 = alarm_wb['Sheet1'].cell(row, 2).value

            full_alarm_xpath = "//*[" + alarm_xpath + str(read_data_col_1) + "'] and " + alarm_xpath + str(
                read_data_col_2) + "']]"
            alarm_display_data = driver.find_element_by_xpath(row_xpath + "[" + str(row + 1) + "]").text
            try:
                driver.find_element_by_xpath(full_alarm_xpath)
            except Exception:
                flag = False
                logger.error("Expected: Alarm Code " + str(read_data_col_1) + ' ' + str(
                    read_data_col_2) + "  - Actual: Alarm Code " + str(replace_data_alarm(alarm_display_data)))
        try:
            logger.info("**********Verify Number of Red Alarm!***************\n")
            num_red_alarm = len(driver.find_elements_by_xpath(red_alarm_xpath))
            print(num_red_alarm)
            if num_red_alarm == 0:
                logger.info("No Red Alarm")
            else:
                logger.error("No. of Red Alarm: " + str(num_red_alarm))
                for t in range(1, num_red_alarm + 1):
                    red_alarm = driver.find_element_by_xpath(red_alarm_xpath + "[" + str(t) + "]").text
                    logger.error("Alarm Code " + str(replace_data_alarm(red_alarm)) + " Status: Red")
        except Exception:
            pass
    except Exception:
        flag = False
    logger.info("**********Alarm information test is completed!***************\n")


def test_diagnostics(browser):
    global flag
    if vr.driver is None:
        br_st.browser_setup(browser)
    logger.info("**********Test_004_DiagnosticsPage***************\n")
    data_diagnostics_verify(vr.driver)
    if flag == False:
        assert False
